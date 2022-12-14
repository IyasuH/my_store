#!/usr/bin/python3
"""
STOCK managment
"""
from kivy.lang import Builder
from kivymd.uix.screen import MDScreen
from kivy.uix.screenmanager import Screen, ScreenManager
from datetime import datetime
from datetime import date
from kivymd.app import MDApp
from kivy.animation import Animation
#from kivymd.tools.hotreload.app import MDApp
from kivymd.uix.floatlayout import MDFloatLayout
from kivymd.uix.tab import MDTabsBase
from kivymd.uix.datatables import MDDataTable
from kivymd.uix.anchorlayout import MDAnchorLayout
from kivymd.uix.button import MDFloatingActionButton, MDRectangleFlatIconButton, MDRaisedButton, MDIconButton, MDFlatButton
from kivymd.uix.label import MDLabel
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.gridlayout import MDGridLayout
from kivymd.uix.filemanager import MDFileManager
from kivymd.uix.list import TwoLineAvatarListItem, ThreeLineAvatarIconListItem, OneLineListItem, TwoLineAvatarIconListItem, OneLineAvatarIconListItem
from kivymd.uix.dialog import MDDialog
from kivymd.toast import toast
from kivy.metrics import dp
from kivy.core.window import Window
import os
from operator import itemgetter
from openpyxl import Workbook
import database
import openpyxl
from openpyxl import load_workbook
from kivy.properties import ListProperty, NumericProperty, StringProperty, ObjectProperty
from kivymd.uix.behaviors import RoundedRectangularElevationBehavior
from kivymd.uix.card import MDCard, MDCardSwipe
from kivy.uix.anchorlayout import AnchorLayout
from kivy.clock import Clock
from kivymd.uix.menu import MDDropdownMenu
from kivymd.uix.textfield.textfield import MDTextField
from kivymd.uix.behaviors import HoverBehavior
from kivymd.theming import ThemableBehavior
from kivymd.uix.list import IRightBodyTouch
from kivy.utils import get_color_from_hex

from kivymd.uix.behaviors import (
    RectangularRippleBehavior,
    BackgroundColorBehavior,
    CommonElevationBehavior,
)

from plyer import notification

def notify(noti_title, noti_msg):
	"""
	this function will be called to handle notifications with title and msg passed as an argument
		noti_title: notification title
		noti_msg: notification message
	Other parametrs also could be added and modified from the default one
	"""
	notification.notify(title=noti_title, message=noti_msg, app_name='My Store', app_icon='', timeout=10, ticker='', toast=False)


# THE NEXT 4 LINES ONLY RUNS ON android device 
# USED TO CALL FOR LOCAL STORAGE IN ANDROID DEVICE

# import android
# from android.permissions import request_permissions, Permission
# from android.storage import primary_external_storage_path
# request_permissions([Permission.WRITE_EXTERNAL_STORAGE, Permission.READ_EXTERNAL_STORAGE])

# for Android 
# primary_ext_storage = primary_external_storage_path()

## And this is save files for PC app 
location=os.getcwd()
primary_ext_storage = location

class MDashCard(MDCard, CommonElevationBehavior):
	"""
	cards on the dashboard
	"""
	winSize = Window.size
	focu_color = ListProperty([1, 1, 1])
	unfocu_color = ListProperty([1, 1, 1])
	text1 = StringProperty("")
	text2 = StringProperty("")
	text3 = StringProperty("")
	# Icon (arrow-top-right, arrow-bottom-right)
	updownIcon = StringProperty("")
	# range b/n 0-100
	cpbValue = NumericProperty(0)
	cpbBarColor = ListProperty([1, 1, 1])

class MDXpense(MDCard):
	"""
	Single expense item card
	For now it is going only to have three values
		Item - Expense name
		Amount - Expense Amount
		Date - date where expense made
	"""
	expenseId = NumericProperty(0)
	expenseName = StringProperty("")
	expenseAmount = StringProperty("")
	expenseDate = StringProperty("")

	def expenseEdit(self):
		"""
		To edit expenses
		"""
		self.dialogEditExpenses = MDDialog(
			title="Edit Expenses info",
			type="custom",
			content_cls=Edit_Expenses_layout(self.expenseId),
			buttons=[
				MDRaisedButton(
					text="Close",
					theme_text_color="Custom",
					on_press=self.closeEditExpenses
				),
				MDRaisedButton(
					text="Delete",
					theme_text_color="Custom",
					on_press=self.deleteExpenses,
					id=str(self.expenseId)
				),

			]
		)
		self.dialogEditExpenses.open()

	def closeEditExpenses(self, inst):
		"""
		close edit Expenses
		"""
		self.dialogEditExpenses.dismiss()

	def deleteExpenses(self, inst):
		"""
		delete specific Expenses
		"""
		expenseId=inst.id
		self.dialogDeleteExpenses = MDDialog(
			title="Are You Sure?",
			text="You won't be able to revert this!",
			#auto_dismiss=False,
			radius=[20, 7, 20, 7],
			buttons=[
				MDFlatButton(
					text="CANCEL",
					theme_text_color="Custom",
					on_press=self.cancleDeleteExpenses
				),
				MDRaisedButton(
					text="DELETE",
					theme_text_color="Custom",
					md_bg_color=(243/255, 63/255, 63/255, 1),
					on_release=self.sureDeleteExpenses,
					id=str(expenseId)
				)
			]
		)
		self.dialogDeleteExpenses.open()
	def cancleDeleteExpenses(self, inst):
		"""
		cancle deleting 
		"""
		self.dialogDeleteExpenses.dismiss()

	def sureDeleteExpenses(self, inst):
		"""
		do the delete
		"""
		expense = database.detailExpenses(inst.id)
		if expense:
			database.deleteExpenses(inst.id)
			toast('Expenses deleted successfully :)')
		else:
			toast('Expenses ID did not found :(')
		self.dialogDeleteExpenses.dismiss()

class MDCashBankCard(MDCard):
	"""
	Single cash bank card
		date - sales date
		bankName - bank Name		
		amount - amount in specified date
		bankId - bankId 
		salesId - slaesId(may use for future expansion like making to see salesItem, customerName, ....)
	"""
	# for now for this class there is no methods to be called
	date = StringProperty("")
	bankName = StringProperty("")
	amount = StringProperty("")
	bankId = NumericProperty(0)
	salesId = NumericProperty(0)
	itemName = StringProperty("")

class MDCashCashCard(MDCard):
	"""
	Single cash cash card
		date - sales date
		bankName - bank Name
		bankId - bankId
		salesId - slaesId(may use for future expansion like making to see salesItem, customerName, ....)
	"""
	date = StringProperty("")
	amount = StringProperty("")
	bankId = NumericProperty(0)
	salesId = NumericProperty(0)
	itemName = StringProperty("")

	def cashToAccount(self):
		"""
		transfer from Cash to Account
			# after cash selected by checker
			# where choosing account will be done by popup
		"""
		# here two queries 
		## first to edit sales bankId since transeferdd from 12 to _
		## second to edit bankAcc to subtract amount from bankId 12 and to add amount to bankId _
		self.selectBankAcc = MDDialog(
			title="Select Bank Account",
			type="confirmation",
			items=[BankConfirm(text=i[1], bankId=str(i[0]))for i in database.readBankAcc()],
			buttons=[
				MDFlatButton(
					text="Cancle",
					on_release=self.selectBankAccCancle
				),
				MDRaisedButton(
					text="Ok",
					on_release=self.selectBankAccOk
				)
			]
		)
		self.selectBankAcc.open()

	def selectBankAccCancle(self, inst):
		"""
		Cancle selectBankAcc Dialog
		"""
		self.selectBankAcc.dismiss()

	def selectBankAccOk(self, inst):
		"""
		for md dialog selectBank OK button callback
		"""
		if bankCID == "":
			toast("please first select bank :(")
			return
		if bankCID == "12":
			toast("That Id is for Cash select another bank :|")
		database.updateBankAmountC(int(self.amount), bankCID, self.bankId)
		"""
		change bankId in sales table from cash 12 to bankCID
			here first argument is salesId
			second argument is bankId it changed to 
		"""
		database.updateBankIdINSalesC(self.salesId, bankCID)
		toast("Cash Transferred Successfully :)")
		self.selectBankAcc.dismiss()


class MDItemCard(MDCard):
	"""
	Single invntory item card
	For now it is going only to have three values
		itemId - Item database Id
		itemAIcon - Item Availability icon just to show the avialbility of item
			using three types of icon(,,) for available, criticalLevel, empty
		itemAIconColor - color for itemAIcon
		itemName - Sales Item Name
		itemQty - Quantity
		itemPrice - single value selling price
		total - amount in money = Qty * amount
	"""
	itemId = NumericProperty(0)
	itemAIcon = StringProperty("")
	itemAIconColor = ListProperty([1,1,1,1])
	itemName = StringProperty("")
	itemQty = StringProperty("")
	itemPrice = StringProperty("")
	total = StringProperty("")
	def item_detail(self, Id):
		"""
		sales detail
		"""
		self.item_detailPopup = MDDialog(
			title="Item detail",
			type="custom",
			content_cls=Detail_item_layout(Id),
			buttons=[
				MDRaisedButton(
					text="Edit",
					on_press=self.editStock,
				),
				MDRaisedButton(
					text="Delete",
					on_press=self.deleteStock,
				)
			]
		)
		self.item_detailPopup.open()

	def editStock(self, inst):
		"""
		Edit specific stock item
		"""
		itemId=self.itemId
		print("inst in ediStock", itemId)
		self.dialogEditItem = MDDialog(
			title="Edit Inventory Item",
			type="custom",
			auto_dismiss=False,
			content_cls=Edit_Item_layout(itemId),
			buttons=[
				MDRaisedButton(
					text="Close",
					theme_text_color="Custom",
					on_press=self.closeEditItem
				),
			]
		)
		self.dialogEditItem.open()
		self.item_detailPopup.dismiss()

	def closeEditItem(self, inst):
		"""
		close edit bank
		"""
		print("close edit")
		self.dialogEditItem.dismiss()

	def deleteStock(self, inst):
		"""
		delete specific stock item
		"""
		itemId=self.itemId
		self.dialogDeleteItem = MDDialog(
			title="Are You Sure?",
			text="You won't be able to revert this!",
			#auto_dismiss=False,
			radius=[20, 7, 20, 7],
			buttons=[
				MDFlatButton(
					text="CANCEL",
					theme_text_color="Custom",
					on_press=self.cancleDeleteItem
				),
				MDRaisedButton(
					text="DELETE",
					theme_text_color="Custom",
					md_bg_color=(243/255, 63/255, 63/255, 1),
					on_release=self.sureDeleteItem,
				)
			]
		)
		self.dialogDeleteItem.open()
		self.item_detailPopup.dismiss()

	def cancleDeleteItem(self, inst):
		"""
		cancle deleting 
		"""
		self.dialogDeleteItem.dismiss()

	def sureDeleteItem(self, inst):
		"""
		do the delete
		"""
		self.dialogDeleteItem.dismiss()
		database.deleteItem(self.itemId)
		toast("Your File Has been deleted.")
		self.dialogDeleteItem.dismiss()


class MDSalesCard(MDCard):
	"""
	Single sales item card
	For now it is going only to have three values
		Item - Sales Item Name
		Qty - Quantity
		amount - amount gained (qty*single value)value
		date - when sales made	
	"""
	salesId = NumericProperty(0)
	salesName = StringProperty("")
	salesQty = StringProperty("")
	salesAmount = StringProperty("")
	salesDate = StringProperty("")
	def sales_detail(self, salesId):
		"""
		sales detail
		"""
		self.sales_detailPopup = MDDialog(
			title="Sales detail",
			type="custom",
			content_cls=Detail_sales_layout(salesId),
			buttons=[
				MDRaisedButton(
					text="Delete",
					on_press=self.deleteSales,
					id = str(salesId),
				)
			]
		)
		self.sales_detailPopup.open()

	def deleteSales(self, inst):
		"""
		delete specific sales
		"""
		salesId=inst.id
		self.dialogDeleteSales = MDDialog(
			title="Are You Sure?",
			text="You won't be able to revert this!",
			#auto_dismiss=False,
			radius=[20, 7, 20, 7],
			buttons=[
				MDFlatButton(
					text="CANCEL",
					theme_text_color="Custom",
					on_press=self.cancleDeleteSales
				),
				MDRaisedButton(
					text="DELETE",
					theme_text_color="Custom",
					md_bg_color=(243/255, 63/255, 63/255, 1),
					on_release=self.sureDeleteSales,
					id=str(salesId)
				)
			]
		)
		self.dialogDeleteSales.open()
		self.sales_detailPopup.dismiss()
	def cancleDeleteSales(self, inst):
		"""
		cancle deleting 
		"""
		self.dialogDeleteSales.dismiss()

	def sureDeleteSales(self, inst):
		"""
		do the delete
		"""
		#####DON'T FORGET THIS####
		#### HERE REQUIRED TO EDIT OTHER TABLES INCLUDING bankAcc(amount), customer(totalsale), item(quantity)####
		sales = database.detailSales(inst.id)
		if sales:
			updateDate = date.today()
			# decreasing customer frequency of purchase and sales revenue
			customer = database.detailCustomer(sales[2])
			database.updateCustomer(customer[0], customer[1], customer[2], customer[3], customer[4], customer[5], customer[6], customer[7]-1, customer[8]-int(sales[5]), customer[9])

			# decreasing bank collected revenue
			bankAccount = database.detailBankAccId(sales[6])
			updateDateStr = updateDate.strftime("%d/%m/%y")
			database.updateBankAcc(bankAccount[0], bankAccount[1], bankAccount[2] - int(sales[5]), bankAccount[3], updateDateStr)

			# increasing item quantity since sales is deleted
			item = database.detailItem(sales[1])
			database.updateItem(item[0], item[1], item[2] + int(sales[3]), item[3], item[4], item[5], updateDateStr)

			# finally deleting sales 
			database.deleteSales(inst.id)
			toast('Sales deleted successfully :)')
		else:
			toast('Sales ID did not found :)')
		self.dialogDeleteSales.dismiss()



class MDCustomCard(MDCard):
	"""
	Crads for Customers page
	"""
	customerId = NumericProperty(0)
	customerName = StringProperty("")
	companyName = StringProperty("")
	customerTinNumber = StringProperty("")
	customerPhoneNumber = StringProperty("")
	totalPurchased = StringProperty(0)
	itemPurchased = ListProperty([])
	def editCustomer(self):
		"""
		edit customer
		"""
		print("on editCustomer")
		self.dialogEditCustomer = MDDialog(
			title="Edit Customer info",
			type="custom",
			content_cls=Edit_Customer_layout(self.customerId),
			buttons=[
				MDRaisedButton(
					text="Close",
					theme_text_color="Custom",
					on_press=self.closeEditCustomer
				)
			]
		)
		self.dialogEditCustomer.open()

	def closeEditCustomer(self, inst):
		"""
		close edit bank
		"""
		print(Edit_Customer_layout(self.customerId).customerName.text)
		self.dialogEditCustomer.dismiss()

	def customerPurchased(self, cId):
		"""
		customer Purchased items
		"""
		self.customerSalesDialog = MDDialog(
			title="Customer Sales INFO",
			type="custom",
			content_cls=Customer_Sales_layout(cId),
			buttons=[
				MDRaisedButton(
					text="Close",
					on_press=self.closeCSales
				),
			]
		)
		self.customerSalesDialog.open()

	def closeCSales(self, inst):
		"""
		cloase Customer Sales
		"""
		self.customerSalesDialog.dismiss()

	def deleteCustomer(self):
		"""
		delete call method
		"""
		print("on deleteCustomer")
		self.dialogDeleteCustomer = MDDialog(
			title="Are You Sure?",
			text="You won't be able to revert this!",
			#auto_dismiss=False,
			radius=[20, 7, 20, 7],
			buttons=[
				MDFlatButton(
					text="CANCEL",
					theme_text_color="Custom",
					on_press=self.cancleDeleteCustomer
				),
				MDRaisedButton(
					text="DELETE",
					theme_text_color="Custom",
					md_bg_color=(243/255, 63/255, 63/255, 1),
					on_release=self.sureDeleteCustomer,
				)
			]
		)
		self.dialogDeleteCustomer.open()
	def cancleDeleteCustomer(self, *args):
		"""
		cancle deleting 
		"""
		self.dialogDeleteCustomer.dismiss()

	def sureDeleteCustomer(self, inst):
		"""
		do the delete
		"""
		self.dialogDeleteCustomer.dismiss()
		database.deleteCustomer(self.customerId)
		print("delete id", self.customerId)
		toast("Your File Has been deleted :)")

class TopClientsListItem(MDCard):
	client_id = NumericProperty(0)
	client_name = StringProperty("")
	company_name = StringProperty("")
	amount = StringProperty("")

	def topCustomerDetail(self):
		"""
		top Customer Detail
		"""
		print("Will outPut Detail customer INFO")

class BarChart(MDFloatLayout):
	"""
	Bar Chart with canvas RoundedRectangle
	Where the python functions are similar only kivy chaged from CircualrProgressBar
	"""
	bar_color = ListProperty([1, 1, 1])
	# where this value is the percent for the progress bar
	set_value = NumericProperty(0) # The increasing value with clock
	text = StringProperty("0") # This the text on the graph and get value from set_value(incresing with clock untill value)
	refer = StringProperty("") # This like Mon, Tue, ..
	value = NumericProperty(0) # final value
	counter = 0 # will count with clock untill == value
	duration = NumericProperty(1.5)
	def __init__(self, **kwargs):
		super(BarChart, self).__init__(**kwargs)
		Clock.schedule_once(self.animate, 0)

	def animate(self, *args):
		if self.value != 0:
			Clock.schedule_interval(self.percent_counter, self.duration/(self.value*200))
		else:
			Clock.schedule_interval(self.percent_counter, 0)
		#Clock.schedule_interval(self.percent_counter, self.duration/(self.value*.))

	def percent_counter(self, *args):
		if self.counter < self.value:
			self.counter += 5
			self.text = f"${self.counter}"
			self.set_value = self.counter
		else:
			Clock.unschedule(self.percent_counter)


class CircularProgressBar(AnchorLayout):
	"""
	Circular progressbar
	"""
	bar_color = ListProperty([1, 1, 1])
	# for PC
	bar_width = NumericProperty(2.5)
	# for Android
	##bar_width = NumericProperty(4)
	# where this value is the percent for the progress bar
	set_value = NumericProperty(0)
	text = StringProperty("0%")
	value = NumericProperty(0)
	counter = 0
	duration = NumericProperty(1.5)

	def __init__(self, **kwargs):
		super(CircularProgressBar, self).__init__(**kwargs)
		Clock.schedule_once(self.animate, 0)

	def animate(self, *args):
		if self.value != 0:
			if self.value < 0:
				Clock.schedule_interval(self.percent_counter_neg, self.duration/self.value)
			elif self.value > 0:
				Clock.schedule_interval(self.percent_counter_pos, self.duration/self.value)
		else:
			Clock.schedule_interval(self.percent_counter_pos, 0)

	def percent_counter_pos(self, *args):
		if self.counter < self.value:
			self.counter += 1 
			self.text = f"+{self.counter}%"
			self.set_value = self.counter			
		else:
			Clock.unschedule(self.percent_counter_pos)

	def percent_counter_neg(self, *args):
		if self.counter > self.value:
			self.counter -= 1 
			self.text = f"{self.counter}%"
			self.set_value = self.counter			
		else:
			Clock.unschedule(self.percent_counter_neg)

# comment this out before deploying
#Window.size = (327, 585)

class Customer_sales_list_item(MDCard):
	"""
	Customer sales list item
	"""
	itemName = StringProperty("")
	salesDate = StringProperty("")
	salesRevenue = StringProperty("")
	def __init__(self, *args, **kwargs):
		super(Customer_sales_list_item, self).__init__(**kwargs)

class Bank_list_item(MDCardSwipe):
	"""
	Bank list items
	"""
	bankName = StringProperty("")
	bankAccountCreatedAt = StringProperty("")
	bankAmount = StringProperty("")
	bankId = NumericProperty(0)
	icon = StringProperty("")
	def __init__(self, *args, **kwargs):
		super(Bank_list_item, self).__init__(**kwargs)
	def deleteAccount(self, bankId):
		"""
		delete call method
		"""
		self.dialogDeleteBank = MDDialog(
			title="Are You Sure?",
			text="You won't be able to revert this!",
			#auto_dismiss=False,
			radius=[20, 7, 20, 7],
			buttons=[
				MDFlatButton(
					text="CANCEL",
					theme_text_color="Custom",
					on_press=self.cancleDeleteBank
				),
				MDRaisedButton(
					text="DELETE",
					theme_text_color="Custom",
					md_bg_color=(243/255, 63/255, 63/255, 1),
					on_release=self.sureDeleteBank
				)
			]
		)
		self.dialogDeleteBank.open()
	def cancleDeleteBank(self, *args):
		"""
		cancle deleting 
		"""
		self.dialogDeleteBank.dismiss()

	def sureDeleteBank(self, *args):
		"""
		do the delete
		"""
		self.dialogDeleteBank.dismiss()
		# THIS TO VOID DELETING CASH ACCOUNT#
		# AND DEPENDINGLY THE bankId MAY CHANGE#
		if self.bankId == 12:
			toast("Cash Account Cannot be deleted :|")
			self.dialogDeleteBank.dismiss()
			return
		else:
			database.deleteBankAcc(self.bankId)
			toast("Your File Has been deleted :)")
			self.dialogDeleteBank.dismiss()

	def editAccount(self, bankId):
		"""
		edit call method
		"""
		#Setting().dismissDialogB()
		"""
			Creates New dialogbox where bank elements
			described in the textfield
		"""
		self.dialogEditBank = MDDialog(
			title="Edit Bank info",
			type="custom",
			auto_dismiss=False,
			content_cls=Edit_Bank_layout(bankId),
			buttons=[
				MDRaisedButton(
					text="Close",
					theme_text_color="Custom",
					on_press=self.closeEditBank
				),
			]
		)
		self.dialogEditBank.open()

	def closeEditBank(self, *args):
		"""
		close edit bank
		"""
		print("close edit")
		self.dialogEditBank.dismiss()

class Bank_list_layout(MDBoxLayout):
	"""

	This class is for listsing bank accounts from database table
	"""
	def __init__(self) -> None:
		super(Bank_list_layout, self).__init__()
		banks = database.readBankAcc()
		self.bank_mdlist.clear_widgets()
		for bank in banks:
			self.bank_mdlist.add_widget(Bank_list_item(
				bankId = bank[0],
				bankName = bank[1],
				bankAccountCreatedAt = bank[3],
				bankAmount = str(bank[2]),
				icon = "bank"
			)
			)

class Edit_Bank_layout(MDBoxLayout):
	"""
	Edit bank info 
		This is going to be called by mddialog in BankList item as content_cls
	"""
	ToDay = date.today()
	def __init__(self, bankId) -> None:
		super(Edit_Bank_layout, self).__init__(bankId)
		self.bankId = bankId

		"""
		This is to change the text of the MDTextField when mddialog open
		"""
		bank = database.detailBankAccId(self.bankId)
		#print(bank)
		self.bankEName.text = bank[1]
		self.bankEAmount.text = str(bank[2])
		self.bankECreatedAt.text = bank[3]

	def saveChanges(self):
		"""
		to save(update) changes to database
		"""
		updatedAt = date.today()
		database.updateBankAcc(self.bankId, self.bankEName.text, self.bankEAmount.text, self.bankECreatedAt.text, updatedAt)
		toast("Data is updated successfully :)")

class Edit_Personal_layout(MDBoxLayout):
	"""
	Edit personal info 
		This is going to be called from setting
	"""
	ToDay = date.today()
	def __init__(self, Id) -> None:
		super(Edit_Personal_layout, self).__init__(Id)
		self.Id = Id

		"""
		This is to change the text of the MDTextField when mddialog open
		"""
		person = database.readSPersonal(self.Id)
		print(person)
		self.personName.text = person[1]

	def saveChanges(self):
		"""
		to save(update) changes to database
		"""
		database.updatePersonalInfo(self.Id, self.personName.text)
		toast("Data is updated successfully :)")


class Edit_Expenses_layout(MDBoxLayout):
	"""
	Edit Expenses
	"""
	def __init__(self, expenseId):
		super(Edit_Expenses_layout, self).__init__(expenseId)
		self.expenseId = expenseId

		exepense = database.detailExpenses(self.expenseId)
		self.expenseType.text = exepense[1]
		self.expenseName.text = exepense[2]
		self.expenseAmount.text = str(exepense[3])
		self.expenseDate.text = exepense[4]

	def saveChanges(self):
		"""
		to save(update) changes to database
		"""
		database.updateExpenses(self.expenseId, self.expenseType.text, self.expenseName.text, self.expenseAmount.text, self.expenseDate.text)
		toast("Data is updated now you can close")


class Edit_Customer_layout(MDBoxLayout):
	"""
	Edit Customer
	"""
	def __init__(self, customerId):
		super(Edit_Customer_layout, self).__init__(customerId)
		self.customerId = customerId

		customer = database.detailCustomer(self.customerId)
		self.customerName.text = customer[1]
		self.companyName.text = customer[2]
		self.tinNumber.text = str(customer[3])
		self.customerCity.text = customer[4]
		self.customerPhoneN.text = str(customer[5])
		self.accountCreatedAt.text = customer[6]
		self.frequencyOfP.text = str(customer[7])
		self.totalP.text = str(customer[8])
		self.accountNumber.text = str(customer[9])

	def saveChanges(self):
		"""
		to save(update) changes to database
		"""
		updatedAt = date.today()
		database.updateCustomer(self.customerId, self.customerName.text, self.companyName.text, self.tinNumber.text, self.customerCity.text, self.customerPhoneN.text, self.accountCreatedAt.text, self.frequencyOfP.text, self.totalP.text, self.accountNumber.text)
		toast("Data is updated now you can close")

class Edit_Item_layout(MDBoxLayout):
	"""
	Edit Inventory Item
	"""
	def __init__(self, itemId):
		super(Edit_Item_layout, self).__init__(itemId)
		self.itemId = itemId

		"""
		This is to change the text of the MDTextField when mddialog open
		"""
		item = database.detailItem(self.itemId)
		self.itemName.text = item[1]
		self.itemQuantity.text = str(item[2])
		self.purchasedAt.text = item[3]
		self.sellingPS.text = str(item[4])
		self.sellingPB.text = str(item[5])

	def saveChanges(self):
		"""
		to save(update) changes to database
		"""
		updatedAt = date.today()
		database.updateItem(self.itemId, self.itemName.text, self.itemQuantity.text, self.purchasedAt.text, self.sellingPS.text, self.sellingPB.text, updatedAt)
		toast("Data is updated now you can close :)")


class New_Bank_layout(MDBoxLayout):
	"""
	Adding New Bank Account
	"""
	ToDay = date.today()
	def __init__(self) -> None:
		super(New_Bank_layout, self).__init__()
	def addNewBank(self):
		"""
		Save Fixed Expense
		"""
		if self.bankName.text and self.bankAmount.text and self.bankCreatedAt.text:
			database.insertBankAcc(str(self.bankName.text), self.bankAmount.text, str(self.bankCreatedAt.text), str(self.bankCreatedAt.text))
			self.bankName.text = ""
			self.bankAmount.text = ""
			self.bankCreatedAt.text = ""
			toast("Data saved now you can close :)")
		else:
			toast("Fill all the info :(")	



class New_Var_Expense_layout(MDBoxLayout):
	"""
	PopUP window to add new variable expenses
	"""
	def __init__(self) -> None:
		super(New_Var_Expense_layout, self).__init__()
	def addNewVarExpense(self):
		"""
		Save Fixed Expense
		"""
		if self.varExpenseItem.text and self.varExpenseAmount.text and self.varExpenseDate.text:
			type = "Variable"
			database.insertXpense(type, str(self.varExpenseItem.text), str(self.varExpenseAmount.text), str(self.varExpenseDate.text))
			self.varExpenseItem.text = ""
			self.varExpenseAmount.text = ""
			self.varExpenseDate.text = ""
			toast("Data saved now you can close :)")
		else:
			toast("Fill all the infor :(")	


class New_FExpense_layout(MDBoxLayout):
	"""
	PopUp window to add new fixed expenses
	"""
	def __init__(self) -> None:
		super(New_FExpense_layout, self).__init__()
	def addNewFExpense(self):
		"""
		Save Fixed Expense
		"""
		if self.fExpenseItem.text and self.fExpenseAmount.text and self.fExpenseDate.text:
			type = "Fixed"
			database.insertXpense(type, str(self.fExpenseItem.text), str(self.fExpenseAmount.text), str(self.fExpenseDate.text))
			self.fExpenseItem.text = ""
			self.fExpenseAmount.text = ""
			self.fExpenseDate.text = ""
			toast("Data saved now you can close :)")
		else:
			toast("Please Fill all the info :(")

class New_customer_layout(MDBoxLayout):
	"""
	PopUp window to add new customer
	"""
	def __init__(self) -> None:
		super(New_customer_layout, self).__init__()

	def addNewCustomer(self):
		"""
		When Add button pressed in the new customer dialog
		"""
		if self.customer_name.text and self.company_name.text and self.customer_tin_number.text and self.customer_city.text and self.customer_phone_number.text and self.customer_account_created_date.text and self.customer_frequency_of_purchase.text and self.customer_total_purchase.text and self.customer_bank_account_number.text:
			database.insertCustomer(self.customer_name.text, self.company_name.text, self.customer_tin_number.text, self.customer_city.text, self.customer_phone_number.text, self.customer_account_created_date.text, self.customer_frequency_of_purchase.text, self.customer_total_purchase.text, self.customer_bank_account_number.text)
			self.customer_name.text = ""
			self.company_name.text = ""
			self.customer_tin_number.text = ""
			self.customer_city.text = ""
			self.customer_phone_number.text = ""
			self.customer_account_created_date.text = ""
			self.customer_frequency_of_purchase.text = ""
			self.customer_total_purchase.text = ""
			self.customer_bank_account_number.text  = ""
			toast('data added successfully now you can close :)')
		else:
			toast('Please fill all info :(')

class New_sales_layout(MDBoxLayout):
	"""
	PopUp window to add new sales
	"""
	def __init__(self) -> None:
		super(New_sales_layout, self).__init__()
		"""
		this is drop down that appers when text field for items touched
		in new sales
		"""
		item_list = database.readSItem()
		item_menu = [
			{
				"viewclass": "OneLineListItem",
				"height": dp(35),
				"text": i[1],
				"max_height": dp(224),
				"on_release": lambda y=i[0]:self.setItem(y),
			}for i in item_list]
		self.listItem = MDDropdownMenu(
			caller = self.item_id,
			items=item_menu,
			position = "bottom",
			width_mult=4,
		)
		"""
		this is drop down that appers when text field for customers touched
		in new sales
		"""
		customer_list = database.readSCustomer()
		customer_menu = [
			{
				"viewclass": "OneLineListItem",
				"height": dp(35),
				"max_height": dp(224),				
				"text": c[1],
				"on_release": lambda y=c[0]:self.setCustomer(y),
			}for c in customer_list]
		self.listCustomer = MDDropdownMenu(
			caller = self.customer_id,
			items=customer_menu,
			position = "bottom",
			width_mult=4,
		)
		"""
		this is drop down that appers when text field for BankAccount touched
		in new sales
		"""
		bank_list = database.readBankAcc()
		bank_menu = [
			{
				"viewclass": "OneLineListItem",
				"height": dp(35),
				"max_height": dp(224),
				"text": b[1],
				"on_release": lambda y=b[0]:self.setBank(y),
			}for b in bank_list]
		self.listBank = MDDropdownMenu(
			caller = self.bank_id,
			items=bank_menu,
			position = "bottom",
			width_mult=4,
		)
		
	def setItem(self, itemId):
		"""
		When item selected from dropdown menu
		"""
		self.item_id.text = str(itemId)
		self.listItem.dismiss()
	def setCustomer(self, customerId):
		"""
		When customer selected from dropdown menu
		"""
		self.customer_id.text = str(customerId)
		self.listCustomer.dismiss()
	def setBank(self, bankId):	
		"""
		When bank selected from dropdown menu
		"""
		self.bank_id.text = str(bankId)
		self.listBank.dismiss()
	
	def calSalesReve(self):
		"""
		calculate sales revenue
		"""
		if self.item_quantitiy.text and self.item_id.text:
			revenu = int(self.item_quantitiy.text) * int(database.detailItem(self.item_id.text)[4])
			self.sales_revenue.text = str(revenu)
			self.sales_revenue.hint_text = "In cherecahro"
		else:
			pass
	def addNewSales(self):
		"""
		When Add button pressed in the new customer dialog
		"""
		if self.item_id.text and self.customer_id.text and self.item_quantitiy.text and self.sold_date.text and self.bank_id.text and self.sales_revenue.text:
			# to update customer frequencyOfPurchase and totalPurchase
			customer = database.detailCustomer(self.customer_id.text)
			if customer:
				database.updateCustomer(int(self.customer_id.text), customer[1], customer[2], customer[3], customer[4], customer[5], customer[6], customer[7]+1, customer[8]+int(self.sales_revenue.text), customer[9])
				toast("Sales Added successfully :)")
			else:
				toast("wrong customer id :(")
				self.customer_id.text = ""
				return
			# to update bankAcc amount and updateAt
			bankAccount = database.detailBankAccId(self.bank_id.text)
			if bankAccount:
				updateDate = date.today()
				updateDateStr = updateDate.strftime("%d/%m/%y")
				database.updateBankAcc(bankAccount[0], bankAccount[1], bankAccount[2]+int(self.sales_revenue.text), bankAccount[3], updateDateStr)
			else:
				toast("wrong bankId id :(")
				self.bank_id.text = ""
				return
			item = database.detailItem(self.item_id.text)
			if item:
				"""
				Here check for item quantity
				"""
				if item[2] < int(self.item_quantitiy.text):
					# no insufficient
					toast("selected item is insufficient :( check Your Inventory!")
					self.item_quantitiy.text  = ""
					self.sales_revenue.text = ""
					return
				else:
					pass
			else:
				toast("Wrong item Id :(")
				self.item_id.text = ""
				return
			updateAt = date.today()
			database.updateItem(item[0], item[1], item[2] - int(self.item_quantitiy.text), item[3], item[4], item[5], updateAt)
			database.insertSales(self.item_id.text, self.customer_id.text, self.item_quantitiy.text, self.sold_date.text, self.sales_revenue.text, self.bank_id.text, self.add_sales_info.text)
			self.item_id.text = ""
			self.customer_id.text = ""
			self.item_quantitiy.text  = ""
			self.sold_date.text  = ""
			self.bank_id.text = ""
			self.sales_revenue.text = ""
			self.add_sales_info.text = ""
			toast('Sales added to database now you can close :)')
		else:
			toast('Please fill all info :(')


class New_stock_layout(MDBoxLayout):
	"""
	Popup to add new item
	"""
	def __init__(self) -> None:
		super(New_stock_layout, self).__init__()

	def addNewItem(self):
		"""
		when new item is added
		"""
		#print("new stoklayout")
		#print(self.item_name.text, self.item_quantity.text, self.purchased_date.text, self.selling_price_single.text, self.selling_price_bulk.text)
		if self.item_name.text and self.item_quantity.text and self.purchased_date.text and self.selling_price_single.text and self.selling_price_bulk.text:
			database.insertItem(self.item_name.text, self.item_quantity.text, self.purchased_date.text, self.selling_price_single.text, self.selling_price_bulk.text)
			toast('Data saved successfully now you can close :)')
			# Then after data entry to avoid data duplication let's clear textBox texts
			self.item_name.text = ""
			self.item_quantity.text = ""
			self.purchased_date.text = ""
			self.selling_price_single.text = ""
			self.selling_price_bulk.text  = ""
		else:
			toast('Please Complete filling :(')

class customerItem(TwoLineAvatarListItem):
	pass

class Detail_sales_layout(MDBoxLayout):
	"""
	sales detail
	"""
	def __init__(self, salesId) -> None:
		super(Detail_sales_layout, self).__init__(salesId)
		#print(Id)
		sales = database.detailSales(salesId)
		self.itemId.secondary_text = str(sales[1])
		self.customerId.secondary_text = str(sales[2])
		self.bankId.secondary_text = str(sales[6])
		self.itemQuantity.secondary_text = str(sales[3])
		self.soldAt.secondary_text = str(sales[4])
		self.salesRevenu.secondary_text = str(sales[5])

class Customer_Sales_layout(MDBoxLayout):
	"""
	Customer Sales 
	"""
	def __init__(self, cId) -> None:
		super(Customer_Sales_layout, self).__init__()
		customerSales = database.readCustomerSales(cId)
		self.customer_sales.clear_widgets()
		for customerSale in customerSales:
			self.customer_sales.add_widget(Customer_sales_list_item(
				itemName = database.detailItem(customerSale[1])[1],
				salesDate = customerSale[4],
				salesRevenue = str(customerSale[5]),
			))


class Detail_item_layout(MDBoxLayout):
	def __init__(self, Id) -> None:
		super(Detail_item_layout, self).__init__()
		#print(Id)
		self.itemId = Id
		self.stock = database.detailItem(Id)
		self.item_name.secondary_text = str(self.stock[1])
		self.item_quantity.secondary_text = str(self.stock[2])
		self.purchased_date.secondary_text = str(self.stock[3])
		self.single_selling_price.secondary_text = str(self.stock[4])
		self.bulk_selling_price.secondary_text = str(self.stock[5])
		self.updated_at.secondary_text = str(self.stock[6])

class BankConfirm(OneLineAvatarIconListItem):
	"""
	This class is to select bankAccount in cash tab
	To transfer money from cash to some bank account
	"""
	divider = None
	bankId = StringProperty("")
	global bankCID
	bankCID = ""
	def set_icon(self, instance_check):
		global bankCID
		if instance_check.active == True:
			bankCID = ""
			instance_check.active = False
		else:
			bankCID = self.bankId			
			instance_check.active = True
		check_list = instance_check.get_widgets(instance_check.group)
		for check in check_list:
			if check != instance_check:
				check.active = False
class Home(MDScreen):
	"""
	main home page
	"""
	#class Stock(MDScreen):
		#"""
		#Stock page
		#"""
	def __init__(self, **kwargs):
		super(Home, self).__init__(**kwargs)

		# file manager when selecting XL file for entering inventory data
		self.item_manager_open = False
		self.item_manager = MDFileManager(
			exit_manager=self.exit_item_manager,
			select_path=self.select_item_path,
		)

		# file manager when selecting XL file for entering customer data
		self.customer_manager_open = False
		self.customer_manager = MDFileManager(
			exit_manager=self.exit_customer_manager,
			select_path=self.select_customer_path,
		)

		#sales
		DATE = datetime.now()
		self.monthForSales = DATE.strftime("%m")
		self.monthShort = DATE.strftime("%b")

		# Dashboard
		DATE = datetime.now()
		self.thisMonth = DATE.strftime("%m")

		#cash
		self.cash_to_bank_list = []
		self.selectedBankAccount = ""

	#=====================DASH-BOARD=====================#
	#def on_tab_dashboard(self):
	def on_enter(self):
		"""
		whene entering to load the first page which is dash-board
		"""
		sales_weekly_data = [[230, "Sun"], [843, "Mon"], [593, "Tue"], [744, "Wed"], [999, "Thu"], [726, "Fri"], [979, "Sat"]] # for each day total of 7
		#profit_weekly_data = [[230, "Sun"], [843, "Mon"], [593, "Tue"], [744, "Wed"], [999, "Thu"], [726, "Fri"], [979, "Sat"]]
		#monthly_data = [] # for each week total of 4
		#yearly_data = [] # for each month total of 12
		#self.salesDashBoardCard.elevation = 0
		#self.salesDashBoardCard.elevation = 4
		#self.profitDashBoardCard.elevation = 0
		#self.customersDashBoardCard.elevation = 4
		#self.expensesDashBoardCard.elevation = 4
		self.helloUserName.text = "Hello "+str(database.readPersonal()[0][1])
		if main.Hr > 6 and main.Hr < 12:
			self.timelyGretting.text = "G o o d   M o r n i n g  !"
		elif main.Hr >12 and main.Hr <18:
			self.timelyGretting.text = "G o o d   A f t e r n o o n  !"
		else:
			self.timelyGretting.text = "G o o d   E v e n i n g  !"
		self.barChartSales.clear_widgets()
		self.barChartProfit.clear_widgets()
		self.topClientsList.clear_widgets()
		############# DASH_BOARD MONTH TO MONTH GROWTH ##############


		#### sales ####
		allSales = database.readSomeSales()
		thisMonthSales = 0
		previousMonthSales = 0
		for x in allSales:
			if int(x[3][3:5]) == int(self.thisMonth):
				thisMonthSales = thisMonthSales + x[2]
			if int(x[3][3:5]) == int(self.thisMonth)-1:
				previousMonthSales = previousMonthSales + x[2]
		if previousMonthSales != 0:
			salesPercentChange = ((thisMonthSales - previousMonthSales)/previousMonthSales)*100
			# text2 is this month sales
			self.salesDashBoardCard.text2 = '$'+str(thisMonthSales)
			# cpbValue is percent change comapre to lastmonth IDK why I called it cpbValue
			self.salesDashBoardCard.cpbValue = salesPercentChange
			if salesPercentChange > 0:
				"""
				to show increasing sine
				"""
				self.salesDashBoardCard.updownIcon = 'arrow-top-right'
			else:
				"""
				to show decreasing sine
				"""
				self.salesDashBoardCard.updownIcon = 'arrow-bottom-right'
			print("Percent change for sales: ", salesPercentChange)
			print("This month profit: ", thisMonthSales)
		else:
			self.salesDashBoardCard.text2 = str(thisMonthSales)
			# if no last month data it will be ZERO(0)
			self.salesDashBoardCard.cpbValue = 0


		#### profit = .25*sales - expenses ####
		# the profit calcualtion is custom
		thisMonthExpense = 0
		previousMonthExpense = 0
		allExpense = database.readExpenses()
		for x in allExpense:
			if int(x[4][3:5]) == int(self.thisMonth):
				thisMonthExpense = thisMonthExpense + x[3]
			if int(x[4][3:5]) == int(self.thisMonth)-1:
				previousMonthExpense = previousMonthExpense + x[3]

		thisMonthProfit = thisMonthSales*0.25 - thisMonthExpense
		previousMonthProfit = previousMonthSales*0.25 - previousMonthExpense
		if previousMonthSales != 0:
			profitPercentChange = ((thisMonthProfit - previousMonthProfit)/previousMonthProfit)*100
			# same priciple as sales 
			self.profitDashBoardCard.text2 = '$'+str(thisMonthProfit)
			self.profitDashBoardCard.cpbValue = profitPercentChange
			if profitPercentChange > 0:
				"""
				to show increasing sine
				"""
				self.profitDashBoardCard.updownIcon = 'arrow-top-right'
			else:
				"""
				to show decreasing sine
				"""
				self.profitDashBoardCard.updownIcon = 'arrow-bottom-right'
			print("Percent change for profit: ", profitPercentChange)
			print("This month profit: ", thisMonthProfit)
		else:
			self.profitDashBoardCard.text2 = str(thisMonthProfit)
			self.profitDashBoardCard.cpbValue = 0


		#### expenses ####
		if previousMonthExpense != 0:
			expensePercentChange = ((thisMonthExpense - previousMonthExpense)/previousMonthExpense)*100
			self.expensesDashBoardCard.text2 = '$'+str(thisMonthExpense)
			self.expensesDashBoardCard.cpbValue = expensePercentChange
			if expensePercentChange > 0:
				"""
				to show increasing sine
				"""
				self.expensesDashBoardCard.updownIcon = 'arrow-top-right'
			else:
				"""
				to show decreasing sine
				"""
				self.expensesDashBoardCard.updownIcon = 'arrow-bottom-right'
			
			print("Percent change for expense:", expensePercentChange)
			print("This month expense", thisMonthExpense)
		else:
			self.expensesDashBoardCard.text2 = str(thisMonthExpense)
			self.expensesDashBoardCard.cpbValue = 0


		#### new customers ####
		thisMonthNewCustomers = 0
		previousMonthNewCustomers = 0
		allCustomers = database.readCustomer()
		for x in allCustomers:
			if int(x[6][3:5]) == int(self.thisMonth):
				thisMonthNewCustomers = thisMonthNewCustomers + 1
			if int(x[6][3:5]) == int(self.thisMonth)-1:
				previousMonthNewCustomers = previousMonthNewCustomers + 1

		if  previousMonthNewCustomers != 0:
			customerPercentChange = ((thisMonthNewCustomers - previousMonthNewCustomers)/previousMonthNewCustomers)*100
			self.customersDashBoardCard.text2 = str(thisMonthNewCustomers)
			self.customersDashBoardCard.cpbValue = customerPercentChange
			if customerPercentChange > 0:
				"""
				to show increasing sine
				"""
				self.customersDashBoardCard.updownIcon = 'arrow-top-right'
			else:
				"""
				to show decreasing sine
				"""
				self.customersDashBoardCard.updownIcon = 'arrow-bottom-right'
			
			print("Percent change for customer", customerPercentChange)
			print("This month customer", thisMonthNewCustomers)
		else:
			self.customersDashBoardCard.text2 = str(thisMonthNewCustomers)
			self.customersDashBoardCard.cpbValue = 0


		sales_weekly_data = [[230, "Sun"], [843, "Mon"], [593, "Tue"], [744, "Wed"], [999, "Thu"], [726, "Fri"], [979, "Sat"]] # for each day total of 7
		sales_weekly_new = [] # may not be 7
		sunSales, monSales, tuesSales, wedSales, thurSales, friSales, satSales = 0,0,0,0,0,0,0
		sunExpen, monExpen, tuesExpen, wedExpen, thurExpen, friExpen, satExpen = 0,0,0,0,0,0,0
		expense_weekly_new = [] # may not be 7
		# then using week number of year to retrive weekly information as barchar

		thisWeekNumber=date.today().strftime("%U") # Sunday first day (00-52)
		for x in allSales:
			# first convert to datetime type 
			salesDate=datetime.strptime(x[3], "%d/%m/%Y")
			# then comapre the week number of each sales with current week number
			if int(salesDate.strftime("%U")) == int(thisWeekNumber):
				# so this is the new sales weekly list
				# then make another kin of list with two data init [salesRvenue, "day"]
				# BUT THIS DOES NOT MEAN I'M GOING TO HAVE ONLY 7 LIST ITEMS
				sales_weekly_new.append([x[2], salesDate.strftime("%a")])
				print(sales_weekly_new)

		for x in allExpense:
			#same procedure as sales
			expenseDate = datetime.strptime(x[4], "%d/%m/%Y")
			if int(expenseDate.strftime("%U")) == int(thisWeekNumber):
				expense_weekly_new.append([x[3], expenseDate.strftime("%a")])
				print(expense_weekly_new)

		for x in expense_weekly_new:
			if x[1] == 'Sun':
				sunExpen = sunExpen + x[0]
			if x[1] == 'Mon':
				monExpen = monExpen + x[0]
			if x[1] == 'Tue':
				tuesExpen = tuesExpen + x[0]
			if x[1] == 'Wed':
				wedExpen = wedExpen + x[0]
			if x[1] == 'Thu':
				thurExpen = thurExpen + x[0]
			if x[1] == 'Fri':
				friExpen = friExpen + x[0]
			if x[1] == 'Sat':
				satExpen = satExpen + x[0]

		for x in sales_weekly_new:
			if x[1] == 'Sun':
				sunSales = sunSales + x[0]
			if x[1] == 'Mon':
				monSales = monSales + x[0]
			if x[1] == 'Tue':
				tuesSales = tuesSales + x[0]
			if x[1] == 'Wed':
				wedSales = wedSales + x[0]
			if x[1] == 'Thu':
				thurSales = thurSales + x[0]
			if x[1] == 'Fri':
				friSales = friSales + x[0]
			if x[1] == 'Sat':
				satSales = satSales + x[0]

		profit_weekly_data = [[(sunSales*.25)-sunExpen, "Sun"], [(monSales*.25)-monExpen, "Mon"], [(tuesSales*.25)-tuesExpen, "Tue"], [(wedSales*.25)-wedExpen, "Wed"], [(thurSales*.25)-thurExpen, "Thu"], [(friSales*.25)-friExpen, "Fri"], [(satSales*.25)-satExpen, "Sat"]]
		print(profit_weekly_data)
		sales_weekly_data = [[sunSales, "Sun"], [monSales, "Mon"], [tuesSales, "Tue"], [wedSales, "Wed"], [thurSales, "Thu"], [friSales, "Fri"], [satSales, "Sat"]] # for each day total of 7
		print(sales_weekly_data)
		#expenses_weekly_data = [[sunExpen, "Sun"], [monExpen, "Mon"], [tuesExpen, "Tue"], [wedExpen, "Wed"], [thurExpen, "Thu"], [friExpen, "Fri"], [satExpen, "Sat"]] # for each day total of 7
		# adding barchar for sales
		[self.barChartSales.add_widget(BarChart(
			value = x[0],
			refer = x[1],
			pos_hint={'center_x': .5, 'center_y': .5},
			size = (57, 140),
			bar_color=[250/255, 115/255, 0]
		))for x in sales_weekly_data]

		# adding barchart for profit
		[self.barChartProfit.add_widget(BarChart(
			value = x[0],
			refer = x[1],
			pos_hint={'center_x': .5, 'center_y': .5},
			size = (57, 140),
			bar_color=[38/255, 255/255, 0]
		))for x in profit_weekly_data]

		####################TOP CLIENTS##############
		allCustomers = database.readSomeCustomer()
		clients = sorted(allCustomers, key = itemgetter(3), reverse = True)[:5]
		[self.topClientsList.add_widget(TopClientsListItem(
			client_id = x[0],
			client_name = x[1],
			company_name = x[2],
			amount = str(x[3])
		))for x in clients]
	#===================================================#

	#=========================STOCK=====================#
	def on_tab_stock(self):
		"""
		when changed to stock tab
		"""
		itemList = database.readSItem()
		i = 0
		totalSA = 0 # total stock value in terms of Birr
		while i < len(itemList):
			if itemList[i][2] > 10:
				iconName = "checkbox-marked-circle"
				iconColor = [39/256, 174/256, 96/256, 1]
			elif itemList[i][2] > 0:
				iconName = "alert"
				iconColor = [255/256, 165/256, 0, 1]
			else:
				iconName = "alert-circle"
				iconColor = [1, 0, 0, 1]
			itemList[i].append(float(itemList[i][2])*float(itemList[i][3]))
			totalSA = totalSA + itemList[i][4]
			itemList[i].append(iconName)
			itemList[i].append(iconColor)
			itemList[i][0] = str(itemList[i][0])
			#itemList[i][1] = "[color=#A7A7A7][b]"+str(itemList[i][1])+"[/b][/color]"
			itemList[i][1] = "[b]"+str(itemList[i][1])+"[/b]"
			itemList[i][2] = "[color=#f3781b][b]"+str(itemList[i][2])+"[/b][/color]"
			itemList[i][3] = "[color=#6d64de][b]$ "+str(itemList[i][3])+"[/b][/color]"
			itemList[i][4] = "[color=#02f363][b]$ "+str(itemList[i][4])+"[/b][/color]"

			#itemList[i][1] = ava
			i += 1

		self.tock_tab.clear_widgets()
		self.totalStockAmount.text = "$ "+str(totalSA)
		[self.tock_tab.add_widget(MDItemCard(
			itemId = x[0],
			itemName = x[1],
			itemQty = str(x[2]),
			itemPrice = str(x[3]),
			total = str(x[4]),
			itemAIcon = str(x[5]),
			itemAIconColor = x[6]
		))for x in itemList]

	def select_item_path(self, path):
		'''
		It will be called when you click on the file name
		or the catalog selection button.
		'''
		self.exit_item_manager()
		# this is for the inventory table
		name, xtension = os.path.splitext(path)
		# first lets check file extension format
		if xtension == '.xlsx':
			print(path)
			wb_obj = load_workbook(path)
			sheet_obj = wb_obj.active
			row = sheet_obj.max_row
			column = sheet_obj.max_column
			if column != 5:
				toast("column number for inventory must be exactly 5 :(")
				return()
			itemFileList=[]
			for i in range(2, row+1):
				tempoTuple = []
				for j in range(1, column+1):
					cell_obj = sheet_obj.cell(row=i, column=j)
					if cell_obj.value == None:
						toast("row: {} and column: {} cell is empty fill them first :(".format(i, j))
						return()
					tempoTuple.append(cell_obj.value)
				itemFileList.append(tempoTuple)
			# here what i done is to change datetime type to str
			# since when uploaded from XL it is datetime type
			for x in itemFileList:
				for a in x:
					if type(a) == datetime:
						x.insert(3, a.strftime("%d/%m/%Y"))
						x.pop(4)
			for itemFi in itemFileList:
				print(itemFi)
				database.insertItem(itemFi[0], itemFi[1], itemFi[2], itemFi[3], itemFi[4])
			toast("Data recorded successfully")
		else: 
			"""
			msg that says invalid extnsion only xlsx accpeted
			"""
			toast("Invalid file type only (.xlsx) type accepted")
			return ()

	def exit_item_manager(self, *args):
		'''
		Called when the user reaches the root of the directory tree
		'''
		self.item_manager_open = False
		self.item_manager.close()

	def uploadStockFromFile(self):
		"""
		uploading stock data from XL file
		"""
		#app_path = os.path.dirname(os.path.abspath(__file__))
		path = os.path.expanduser("~")
		#androidPath = os.path.join(os.getenv('EXTERNAL_STORAGE'), 'Downloads'
		self.item_manager.show(path)
		## For Android
		#self.item_manager.show(primary_ext_storage)

		self.item_manager_open = True
	def newStock(self):
		"""
		new individual stock
		"""
		self.dialogNewStock = MDDialog(
			title="New Item",
			type="custom",
			auto_dismiss=False,
			content_cls=New_stock_layout(),
			buttons=[
				MDRaisedButton(
					text="Close",
					theme_text_color="Custom",
					on_press=self.closeNewStock
				)
			]
		)
		self.dialogNewStock.open()
	def closeNewStock(self, inst):
		#self.stock_tables.update_row_data(inst, itemList)
		self.on_tab_stock()
		self.dialogNewStock.dismiss()

	#===================================================#

	#=========================SALES=====================#

	def on_tab_sales(self):
		"""
		"""
		self.monthDropDownMenuCaller.text = str(self.monthShort)
		months=(("Jan", 1), ("Feb", 2), ("Mar", 3), ("Apr", 4), ("May", 5), ("Jun", 6), ("Jul", 7), ("Aug", 8), ("Sep", 9), ("Oct", 10), ("Nov", 11), ("Dec", 12))
		monthMenu_items = [
			{
				"viewclass": "OneLineListItem",
				"text": f"{i[0]}",
				"bold": True,
				"font_size": "20sp",
				"height": dp(33),
				"on_release": lambda x=i: self.changeMonthSales(x)
			}for i in months
		]
		self.main_monthMenu = MDDropdownMenu(
			items=monthMenu_items,
			caller=self.monthDropDownMenuCaller,
			width_mult=1.3,
		)

		self.salesTab.clear_widgets()
		# here to made filter based on month
		allSales = database.readSomeSales()
		monthSales = []
		for x in allSales:
			if int(x[3][3:5]) == int(self.monthForSales):
				monthSales.append(x)
		# convert (allSales) to (monthSales)
		# here what i do is to convert itemId to itemName
		totalSales = 0
		for x in monthSales:
			x[0] = database.detailItem(x[0])[1]
			totalSales = totalSales + x[2]

		self.totalSalesAmount.text = "$ "+str(totalSales)
		[self.salesTab.add_widget(MDSalesCard(
			salesId = str(x[4]),
			salesName = str(x[0]),
			salesQty = str(x[1]),
			salesAmount = "$ "+str(x[2]),
			salesDate = str(x[3])
		))for x in monthSales]

	def changeMonthSales(self, selMonth):
		"""
		When month selected
		"""
		print(selMonth)
		#self.monthDropDownMenuCaller.text=selMonth[0]
		self.monthForSales = selMonth[1]
		self.monthShort = selMonth[0]
		self.on_tab_sales()

	def reFresh(self):
		"""
		to refesh changes on the database to screen
			currently in used in the customer bar
		"""
		self.on_tab_sales()

	def exportSales(self):
		print("export sales func")
		"""
		exporting sales data
		"""
		allSales = database.readSales()
		monthSales = []
		for x in allSales:
			if int(x[4][3:5]) == int(self.monthForSales):
				monthSales.append(x)
		book = Workbook()
		sheet = book.active
		col0 =['Sales Id', 'Item Id', 'Customer Id', 'Item Quantity', 'Sold Date', 'Sales Revenue', 'Bank Id']

		monthSales.insert(0, col0)
		for sales in monthSales:
			sheet.append(sales)
		#androidPath = os.path.join(os.getenv('EXTERNAL_STORAGE'), 'Downloads')
		book.close()
		book.save("{}/{}_{}_{}_sales.xlsx".format(primary_ext_storage, self.monthShort, date.today().strftime("%Y"), date.today().strftime("%d")))
		#book.close()
		toast("Your file is saved as {}_{}_sales.xlsx".format(self.monthShort, date.today().strftime("%Y")))

	def newSales(self):
		"""
		Adding new single sales to database
		"""
		self.dialogNewSales = MDDialog(
			title="New Sales",
			type="custom",
			auto_dismiss=False,
			content_cls=New_sales_layout(),
			buttons=[
				MDRaisedButton(
					text="Close",
					theme_text_color="Custom",
					on_press=self.closeNewSales
				),
			]
		)
		self.dialogNewSales.open()

	def closeNewSales(self, inst):
		"""
		close button in add new sales dialog
		"""
		self.dialogNewSales.dismiss()
		self.on_tab_sales()

	#===================================================#

	#=========================CASH=====================#
	def on_tab_cash(self):
		"""
		on tab tab_cash
		"""
		namSize = 12
		totalInCash = 0
		totalInBank = 0
		self.cashBank_tab.clear_widgets()
		allSales = database.readCashSales()
		cash = []
		bank=[]
		for x in allSales:
			if x[2] == 12:
				totalInCash = totalInCash + x[1]
				x[0] = "[color=#02f363][b]"+x[0]+"[/b][/color]"
				x[2] = str(x[2])
				x[3] = str(x[3])
				x.insert(4, str(database.detailItem(x[4])[1])) # this is for item name
				cash.append(x)
			else:
				totalInBank = totalInBank + x[1]
				x[0] = "[color=#02f363][b]"+x[0]+"[/b][/color]"
				bank.append(x)
		for x in bank:
			bankName = database.detailBankAccId(x[2])
			x.insert(1, bankName[1])
			x.append(str(database.detailItem(x[5])[1]))
			x[1] = "[color=#f3781b][b]"+str(x[1])+"[/b][/color]"
			x[2] = "[color=#6d64de][b]$ "+str(x[2])+"[/b][/color]"
			#x[3] = "[size=17][color=#a7a7a7b3]"+str(x[3])+"[/color][/size]"
			#x[4] = "[size=17][color=#a7a7a7b3]"+str(x[4])+"[/color][/size]"
		print(cash)
		print(bank)
		self.cashCash_tab.clear_widgets()
		self.cashBank_tab.clear_widgets()
		self.totalInCashAmount.text = "$ "+str(totalInCash)
		self.totalInBankAmount.text = "$ "+str(totalInBank)
		[self.cashBank_tab.add_widget(MDCashBankCard(
			date = x[0],
			bankName = x[1],
			amount = x[2],
			bankId = x[3],
			salesId = x[4],
			itemName = (str(x[6]))
		))for x in bank]

		[self.cashCash_tab.add_widget(MDCashCashCard(
			date = x[0],
			amount = str(x[1]),
			bankId = x[2],
			salesId = x[3],
			itemName = str(x[4])
		))for x in cash]

	# def cashCash_check_press(self, instance_table, current_row):
	# 	"""
	# 	when check pressed for cashCash table
	# 	"""
	# 	activeRows = instance_table.get_row_checks()
	# 	bank_sales = []
	# 	"""
	# 	bank sales have three componenets
	# 		0 = sales Id to change bankId
	# 		1 = sales revenue to edit bankAcc
	# 		2 = cash Id(if cash changes Id changes in the future)
	# 	"""
	# 	for x in activeRows:
	# 		bank_sales.append([x[3], x[1], x[2]])
	# 	print(bank_sales)
	# 	self.cash_to_bank_list = bank_sales

	#===================================================#

	#=========================CUSTOMER=====================#


	def on_tab_customer(self):
		"""
		"""
		self.customerTab.clear_widgets()
		customersList = database.readSCustomer()
		[self.customerTab.add_widget(MDCustomCard(
			customerId=x[0],
			customerName = x[1],
			companyName = x[2],
			customerTinNumber = str(x[5]),
			customerPhoneNumber = str(x[4]),
			totalPurchased = str(x[6]),
			#itemPurchased = database.read,
		))for x in customersList]

	def newCustomer(self):
		"""
		Adding new single customer to database
		"""
		print('on new customers')
		self.dialogNewCustomer = MDDialog(
			title="New Customer",
			type="custom",
			auto_dismiss=False,
			content_cls=New_customer_layout(),
			buttons=[
				MDRaisedButton(
					text="Close",
					theme_text_color="Custom",
					on_press=self.closeNewCustomer
				),
			]
		)
		self.dialogNewCustomer.open()
	def closeNewCustomer(self, inst):
		"""
		close button in add new customer dialog
		"""
		self.dialogNewCustomer.dismiss()

	def uploadCustomerFromFile(self):
		"""
		uploading customer data from XL file
		"""
		path = os.path.expanduser("~")
		self.customer_manager.show(path)

		## Android
		#self.customer_manager.show(primary_ext_storage)
		self.customer_manager_open = True

	def select_customer_path(self, path):
		'''
		It will be called when you click on the file name
		or the catalog selection button.
		'''
		self.exit_customer_manager()
		# this is for the inventory table
		name, xtension = os.path.splitext(path)
		# first lets check file extension format
		if xtension == '.xlsx':
			wb_obj = load_workbook(path)
			sheet_obj = wb_obj.active
			row = sheet_obj.max_row
			column = sheet_obj.max_column
			if column != 9:
				toast("column number for customer must be exactly 9 :(")
				return()
			customerFileList=[]
			for i in range(2, row+1):
				tempoTuple = []
				for j in range(1, column+1):
					cell_obj = sheet_obj.cell(row=i, column=j)
					if cell_obj.value == None:
						toast("row: {} and column: {} cell is empty fill them first :(".format(i, j))
						return()
					tempoTuple.append(cell_obj.value)
				customerFileList.append(tempoTuple)
			# here what i done is to change datetime type to str
			# since when uploaded from XL it is datetime type
			for x in customerFileList:
				for a in x:
					if type(a) == datetime:
						x.insert(5, a.strftime("%d/%m/%Y"))
						x.pop(6)
			for customerFi in customerFileList:
				print(customerFi)
				database.insertCustomer(customerFi[0], customerFi[1], customerFi[2], customerFi[3], customerFi[4], customerFi[5], customerFi[6], customerFi[7], customerFi[8])
			toast("Data recorded successfully")
		else: 
			"""
			msg that says invalid extnsion only xlsx accpeted
			"""
			toast("Invalid file type only (.xlsx) type accepted")
			return ()

	def exit_customer_manager(self, *args):
		'''
		Called when the user reaches the root of the directory tree
		'''
		self.customer_manager_open = False
		self.customer_manager.close()			


	def reFresh(self):
		"""
		to refesh changes on the database to screen
			currently in used in the customer bar
		"""
		self.on_tab_customer()

	#===================================================#

class NewUserInfo(MDScreen):
	"""
	New Account Page
	"""
	def createNewUser(self, userName):
		"""
		To Validate and Add User info to database
		"""
		if userName != "":
			database.CreateUser(userName, main.ToDay)
			return "created"

class Expneses(MDScreen):
	"""
	Expense page
	"""
	def __init__(self, **kwargs):
		"""
		Drop down monthMenu
		"""
		super(Expneses, self).__init__(**kwargs)
		DATE = datetime.now()
		self.monthForExpenses = DATE.strftime("%m")
		self.monthShort = DATE.strftime("%b")

	def changeMonthExpenses(self, selMonth):
		"""
		When month selected
		"""
		self.monthForExpenses = selMonth[1]
		self.monthShort = selMonth[0]
		self.on_enter()

	def on_enter(self):
		"""
		Whene entered to expenses
			# select name,amount,date from expenses (also check by type so type too)
		"""
		self.monthDropDownMenuCaller.text=str(self.monthShort)
		months=(("Jan", 1), ("Feb", 2), ("Mar", 3), ("Apr", 4), ("May", 5), ("Jun", 6), ("Jul", 7), ("Aug", 8), ("Sep", 9), ("Oct", 10), ("Nov", 11), ("Dec", 12))
		monthMenu_items = [
			{
				"viewclass": "OneLineListItem",
				"bold": True,
				"font_size": "20sp",
				"text": f"{i[0]}",
				"height": dp(33),
				"on_release": lambda x=i: self.changeMonthExpenses(x)
			}for i in months
		]
		self.main_monthMenu = MDDropdownMenu(
			items=monthMenu_items,
			caller=self.monthDropDownMenuCaller,
			width_mult=1.3,
		)

		self.fixedExpenses.clear_widgets()
		self.variableExpenses.clear_widgets()
		# here to made filter based on month
		fixExpense = database.readSomeFixExpenses
		varExpense = database.readSomeVarExpenses
		monthFixed = []
		monthVariable = []
		totalVariableExpenses = 0
		totalFixedExpenses = 0
		for x in fixExpense():
			if int(x[3][3:5]) == int(self.monthForExpenses):
				monthFixed.append(x)
				totalFixedExpenses = totalFixedExpenses + x[2]
		for x in varExpense():
			if int(x[3][3:5]) == int(self.monthForExpenses):
				monthVariable.append(x)
				totalVariableExpenses = totalVariableExpenses + x[2]

		self.totalVariableExpenses.text = "$ "+str(totalVariableExpenses)
		self.totalFixedExpenses.text = "$ "+str(totalFixedExpenses)
		self.totalExpenses.text = "$ "+str(totalVariableExpenses + totalFixedExpenses)
		[self.fixedExpenses.add_widget(MDXpense(
			expenseId = x[0],
			expenseName = "[color=#a7a7a7b3][b]"+str(x[1])+"[/color][/b]",
			expenseAmount = "[color=#ff005a][b]$ "+str(x[2])+"[/color][/b]",
			expenseDate = "[color=#02f363][b]"+str(x[3])+"[/color][/b]"
		))for x in monthFixed]

		[self.variableExpenses.add_widget(MDXpense(
			expenseId = x[0],
			expenseName = "[color=#a7a7a7b3][b]"+str(x[1])+"[/color][/b]",
			expenseAmount = "[color=#ff005a][b]$ "+str(x[2])+"[/color][/b]",
			expenseDate = "[color=#02f363][b]"+str(x[3])+"[/color][/b]"
		))for x in monthVariable]
		##################

	def addFixedE(self):
		"""
		Add Fixed Expense
		so here we mddialog where other expense details filled
			except for the type that since that is fixed
		"""
		self.dialogNewFExpense = MDDialog(
			title="New Fixed Expense",
			type="custom",
			auto_dismiss=False,
			content_cls=New_FExpense_layout(),
			buttons=[
				MDRaisedButton(
					text="Close",
					theme_text_color="Custom",
					on_press=self.closeNewFExpense
				),
			]
		)
		self.dialogNewFExpense.open()

	def addVarE(self):
		"""
		Add Fixed Expense
		so here we mddialog where other expense details filled
			except for the type that since that is fixed
		"""
		self.dialogNewVarExpense = MDDialog(
			title="New Variable Expense",
			type="custom",
			auto_dismiss=False,
			content_cls=New_Var_Expense_layout(),
			buttons=[
				MDRaisedButton(
					text="Close",
					theme_text_color="Custom",
					on_press=self.closeNewVarExpense
				),
			]
		)
		self.dialogNewVarExpense.open()


	def closeNewFExpense(self, inst):
		self.dialogNewFExpense.dismiss()
		self.on_enter()

	def closeNewVarExpense(self, inst):
		self.dialogNewVarExpense.dismiss()
		self.on_enter()

	def exportExpenses(self):
		"""
		exporting exepense data from dataTable for specifc Month
		"""
		allExpense = database.readExpenses
		monthExpense = []

		for x in allExpense():
			if int(x[4][3:5]) == int(self.monthForExpenses):
				monthExpense.append(x)

		book = Workbook()
		sheet = book.active
		col0 = ['Expenses Id', 'Expenses Type', 'Expenses Name', 'Amount', 'Date']

		monthExpense.insert(0, col0)
		for expenses in monthExpense:
			sheet.append(expenses)
		book.close()
		book.save("{}/{}_{}_expenses.xlsx".format(primary_ext_storage, self.monthShort, date.today().strftime("%Y")))
		#book.close()
		toast("Your file is saved as {}_{}_sales.xlsx".format(self.monthShort, date.today().strftime("%Y")))


class Setting(MDScreen):
	"""
	Settings page
	"""
	#def on_enter(self):
		#if main().theme_cls.theme_style == "Dark":
			#self.nightModeSwitch.active = True
		#else:
			#self.nightModeSwitch.active = False
	def on_enter(self):
		self.accNameEmailNumber.text = str(database.readPersonal()[0][1])
		self.personId = database.readPersonal()[0][0]
	def edit_personal_info(self, personId):
		"""
		edit call method
		"""
		self.dialogEditPersonal = MDDialog(
			title="Edit Personal info",
			type="custom",
			auto_dismiss=False,
			content_cls=Edit_Personal_layout(personId),
			buttons=[
				MDRaisedButton(
					text="Close",
					theme_text_color="Custom",
					on_press=self.closeEditPersonal
				),
			]
		)
		self.dialogEditPersonal.open()

	def closeEditPersonal(self, *args):
		"""
		close edit bank
		"""
		print("close edit")
		self.dialogEditPersonal.dismiss()


	def bankList(self):
		self.dialogBankList = MDDialog(
			title="Bank list",
			type="custom",
			content_cls=Bank_list_layout(),
			buttons=[
				MDRaisedButton(
					text="Close",
					theme_text_color="Custom",
					on_press=self.closeBankList
				),
				MDRaisedButton(
					text="New",
					theme_text_color="Custom",
					on_press=self.newBank
				),

			]
		)

		self.dialogBankList.open()

	def closeBankList(self, *args):
		print("dismiss")
		self.dialogBankList.dismiss(force=True)

	def newBank(self, inst):
		"""
		New Bank MDDialog
		"""
		self.dialogBankList.dismiss()
		self.dialogNewBank = MDDialog(
			title="New Bank",
			type="custom",
			auto_dismiss=False,
			content_cls=New_Bank_layout(),
			buttons=[
				MDRaisedButton(
					text="Close",
					theme_text_color="Custom",
					on_press=self.closeNewBank
				),
			]
		)
		self.dialogNewBank.open()


	def closeNewBank(self, inst):
		self.dialogNewBank.dismiss()

	def changeNightMode(self):
		"""
		This method is to turn ON/OFF dark mode from setting
		"""
		if self.nightModeSwitch.active == True:
			"""
			was darkMode now to lightMode
			"""
			main().theme_cls.theme_style = "Light"
			print("was darkMode now to lightMode")
			self.nightModeSwitch.active = False
		else:
			"""
			was lightMode now to darkMode
			"""
			main().theme_cls.theme_style = "Dark"
			print("was lightMode now to darkMode")
			self.nightModeSwitch.active = True

	def deleteEveryThing(self):
		"""
		deleteEveryThing call method
		"""
		self.dialogDeleteEveryThings = MDDialog(
			title="Are You Sure?",
			text="This will delete everything on the database including\n(customers, inventory, sales, exepenses and bank account data)\n You won't be able to revert this!",
			#auto_dismiss=False,
			radius=[20, 7, 20, 7],
			buttons=[
				MDFlatButton(
					text="CANCEL",
					theme_text_color="Custom",
					on_press=self.cancleDeleteEveryThing
				),
				MDRaisedButton(
					text="DELETE",
					theme_text_color="Custom",
					md_bg_color=(243/255, 63/255, 63/255, 1),
					on_release=self.sureDeleteEveryThing
				)
			]
		)
		self.dialogDeleteEveryThings.open()
	def cancleDeleteEveryThing(self, *args):
		"""
		cancle deleting 
		"""
		self.dialogDeleteEveryThings.dismiss()

	def sureDeleteEveryThing(self, *args):
		"""
		to delete every thing from database
		"""
		self.dialogDeleteEveryThings.dismiss()
		print('[INFO]: Now DataBase is Wiped')
		database.deleteEveryThingFromDatabase()


class main(MDApp):
	"""
	main class
	"""
	overlay_color = get_color_from_hex("#6042e4")
	currentTime = datetime.now()
	Hr = currentTime.hour
	ToDay = date.today()
	def build(self, first=False):
		"""
		build method
		"""
		if self.Hr < 18:
			self.theme_cls.theme_style = "Dark"
		else:
			self.theme_cls.theme_style = "Dark"
		self.theme_cls.primary_palette = 'Blue'
		self.theme_cls.material_style ="M3"
		screen_manager = ScreenManager()
		if database.readPersonal() == []:
			"""
			If no personal information added to the to call
				page that lets user enter personal info
				most likely when new person install it
			But if there is some personal info no need to call that
			"""
			screen_manager.add_widget(Builder.load_file("new_user_info.kv"))
		screen_manager.add_widget(Builder.load_file("home.kv"))
		screen_manager.add_widget(Builder.load_file("setings.kv"))
		screen_manager.add_widget(Builder.load_file("expenses.kv"))
		return screen_manager

	def backHome(self):
		"""
		This method is to go back to home
		"""
		self.root.transition.direction = "right"
		self.root.current = "home"

	def gotoSetting(self):
		"""
		to goto settings
		"""
		print('settings')
		self.root.transition.direction = "left"
		self.root.current = "setting"
	
	def nightModeSwicthActive(self):
		print("on active")
		self.theme_cls.theme_style = "Dark"

	def nightModeSwicthState(self):
		print("on state")
		self.theme_cls.theme_style = "Light"		

if __name__ == "__main__":
	main().run()
